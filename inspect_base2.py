import openpyxl
import datetime

LOCAL_EXCEL = r"C:\Users\sup.luciana\Desktop\AntiGravity\BANCO DE HORAS\TEMPOS_2026.xlsx"
wb = openpyxl.load_workbook(LOCAL_EXCEL, data_only=True)
ws = wb['BASE']

count = 0
for row_idx in range(2, ws.max_row + 1):
    val = ws.cell(row=row_idx, column=1).value
    if isinstance(val, datetime.datetime) and val < datetime.datetime(2026, 3, 16):
        count += 1

print(f"Rows before March 16: {count}")
