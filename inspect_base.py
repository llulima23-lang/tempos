import openpyxl
import datetime

LOCAL_EXCEL = r"C:\Users\sup.luciana\Desktop\AntiGravity\BANCO DE HORAS\TEMPOS_2026.xlsx"
wb = openpyxl.load_workbook(LOCAL_EXCEL, data_only=True)
ws = wb['BASE']

min_date = None
for row_idx in range(2, ws.max_row + 1):
    val = ws.cell(row=row_idx, column=1).value
    if isinstance(val, datetime.datetime):
        if min_date is None or val < min_date:
            min_date = val

print(f"Min date in BASE: {min_date}")
