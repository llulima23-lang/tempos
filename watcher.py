import time
import os
import shutil
import json
import openpyxl
import datetime

SOURCE_EXCEL = r"C:\Users\sup.luciana\Meu Drive\MF\MF\Indicadores de Cobrança\TEMPOS 2026.xlsx"
LOCAL_EXCEL = r"C:\Users\sup.luciana\Desktop\AntiGravity\BANCO DE HORAS\TEMPOS_2026.xlsx"
JS_OUTPUT = r"C:\Users\sup.luciana\Desktop\AntiGravity\BANCO DE HORAS\data_embedded.js"
JSON_OUTPUT = r"C:\Users\sup.luciana\Desktop\AntiGravity\BANCO DE HORAS\data.json"

def time_to_seconds(val):
    if val is None: return 0
    if isinstance(val, datetime.timedelta): return int(val.total_seconds())
    if isinstance(val, datetime.time): return val.hour * 3600 + val.minute * 60 + val.second
    if isinstance(val, datetime.datetime): return val.hour * 3600 + val.minute * 60 + val.second
    return 0

def process_excel():
    print(f"[{datetime.datetime.now()}] Processando nova versão do Excel...")
    try:
        adm_file = r"C:\Users\sup.luciana\Meu Drive\MF\MF\Indicadores de Cobrança\ADM EQUIPES.xlsx"
        admissao_map = {}
        if os.path.exists(adm_file):
            try:
                wb_adm = openpyxl.load_workbook(adm_file, data_only=True)
                if 'ESPELHO' in wb_adm.sheetnames:
                    ws_adm = wb_adm['ESPELHO']
                    col_mat, col_adm = None, None
                    for c in range(1, ws_adm.max_column + 1):
                        val = str(ws_adm.cell(row=1, column=c).value).strip().lower()
                        if 'matricula' in val or 'matrícula' in val: col_mat = c
                        if 'admissão' in val or 'admissao' in val: col_adm = c
                    if col_mat and col_adm:
                        for r in range(2, ws_adm.max_row + 1):
                            mat = ws_adm.cell(row=r, column=col_mat).value
                            adm = ws_adm.cell(row=r, column=col_adm).value
                            if mat is not None and isinstance(adm, datetime.datetime):
                                admissao_map[str(mat).strip()] = adm.strftime("%Y-%m-%d")
            except Exception as e:
                print(f"Erro ao ler ADM EQUIPES: {e}")

        shutil.copy2(SOURCE_EXCEL, LOCAL_EXCEL)
        wb = openpyxl.load_workbook(LOCAL_EXCEL, data_only=True)
        ws = wb['BASE']
        target_date = datetime.datetime(2026, 3, 16)
        
        records = []
        for row_idx in range(2, ws.max_row + 1):
            date_val = ws.cell(row=row_idx, column=1).value
            if not isinstance(date_val, datetime.datetime): continue
            if date_val < target_date: continue
            
            record = {
                "data": date_val.strftime("%Y-%m-%d"),
                "ano": ws.cell(row=row_idx, column=2).value,
                "mes": ws.cell(row=row_idx, column=3).value,
                "agente": ws.cell(row=row_idx, column=4).value,
                "grupo": ws.cell(row=row_idx, column=5).value,
                "pausas_total": time_to_seconds(ws.cell(row=row_idx, column=6).value),
                "primeiro_login": time_to_seconds(ws.cell(row=row_idx, column=7).value),
                "ultimo_logout": time_to_seconds(ws.cell(row=row_idx, column=8).value),
                "tempo_logado": time_to_seconds(ws.cell(row=row_idx, column=9).value),
                "meta": time_to_seconds(ws.cell(row=row_idx, column=10).value),
                "credito": time_to_seconds(ws.cell(row=row_idx, column=11).value),
                "deficit": time_to_seconds(ws.cell(row=row_idx, column=12).value),
            }
            records.append(record)
            
        ws_resumo = wb['RESUMO']
        resumo = []
        for row_idx in range(2, ws_resumo.max_row + 1):
            nome = ws_resumo.cell(row=row_idx, column=1).value
            if nome and not str(nome).startswith('Atualizado') and not str(nome).startswith('A coluna') and not str(nome).startswith('Pagamento'):
                matricula_val = ws_resumo.cell(row=row_idx, column=8).value
                admissao_val = None
                if matricula_val is not None:
                    admissao_val = admissao_map.get(str(matricula_val).strip())
                entry = {
                    "nome": str(nome),
                    "matricula": matricula_val,
                    "admissao": admissao_val,
                    "operacao": ws_resumo.cell(row=row_idx, column=9).value,
                    "banco_horas": time_to_seconds(ws_resumo.cell(row=row_idx, column=4).value),
                }
                resumo.append(entry)
                
        output = {"records": records, "resumo": resumo, "updated_at": datetime.datetime.now().isoformat()}
        
        # Save as JSON
        with open(JSON_OUTPUT, 'w', encoding='utf-8') as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
            
        # Save as JS Embedded
        js_content = f"const EMBEDDED_DATA = {json.dumps(output, ensure_ascii=False)};\n"
        with open(JS_OUTPUT, 'w', encoding='utf-8') as f:
            f.write(js_content)
            
        print("Sistema atualizado com sucesso!")
    except Exception as e:
        print(f"Erro ao processar: {e}")

def watch_file():
    if not os.path.exists(SOURCE_EXCEL):
        print(f"Arquivo não encontrado: {SOURCE_EXCEL}")
        return
        
    last_mtime = os.path.getmtime(SOURCE_EXCEL)
    print(f"Iniciando monitoramento do arquivo: {SOURCE_EXCEL}")
    
    while True:
        time.sleep(10) # Verifica a cada 10 segundos
        try:
            current_mtime = os.path.getmtime(SOURCE_EXCEL)
            if current_mtime != last_mtime:
                last_mtime = current_mtime
                print("Arquivo Excel modificado! Atualizando base de dados...")
                time.sleep(2) # Espera 2s para garantir que o Excel terminou de salvar
                process_excel()
        except Exception as e:
            pass # Ignora erros de leitura (ex: arquivo em uso)

if __name__ == "__main__":
    process_excel()
    watch_file()
